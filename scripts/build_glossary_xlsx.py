import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAW = Path(__file__).parent.parent / "data" / "raw"
OUT = Path(__file__).parent.parent / "docs" / "quranic-structural-glossary.xlsx"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="33564F", end_color="33564F")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="33564F")
NOTE_FONT = Font(name=FONT, italic=True, size=9, color="6B6558")
BODY_FONT = Font(name=FONT, size=11)
ARABIC_FONT = Font(name=FONT, size=12)
THIN = Side(style="thin", color="D8D0BD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load(name):
    return json.loads((RAW / name).read_text(encoding="utf-8"))


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_body_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right" if c <= 2 else "left", vertical="center", wrap_text=True)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---- Sheet 1: Glossary (term / meaning / category) -------------------------
# Structure mirrors the community's own "جدول حصر المصطلحات القرآنية" sheet
# (مصطلح | شرح المعنى | تصنيف) so rows can be pasted directly into it.

glossary = [
    ("الْجُزْء", "Juz'", "One of 30 equal-length divisions of the Qur'an's text, used to pace a complete recitation over a month.", "هيكلة المصحف", 30),
    ("الْحِزْب", "Hizb", "One of 60 divisions (two per juz'), each further split into four quarters (arba').", "هيكلة المصحف", 60),
    ("الرُّبُع", "Rub' al-Hizb", "A quarter-hizb; the finest standard fixed subdivision of the Mushaf, 240 in total.", "هيكلة المصحف", 240),
    ("الْمَنْزِل", "Manzil", "One of 7 divisions used to complete the Qur'an's recitation over a week.", "هيكلة المصحف", 7),
    ("السُّورَة", "Surah", "A chapter of the Qur'an; 114 in total, each with a name, a place of revelation (Meccan/Medinan), and a fixed ayah count.", "هيكلة المصحف", 114),
    ("الْآيَة", "Ayah", "A verse; the basic textual unit of a surah. Numbering is a scholarly convention (ʿadd al-āy) that can differ slightly between transmission traditions.", "هيكلة المصحف", 6236),
    ("الرُّكُوع", "Ruku'", "A thematic passage marker used in some Mushaf printings (mainly Indo-Pak) to mark natural pause points within a surah.", "هيكلة المصحف", 558),
    ("السَّجْدَة", "Sajdah (tilawah)", "A verse containing a prostration point; reciting or hearing it is traditionally followed by a physical prostration. 15 in the standard Hafs list, marked 'required' or 'optional' by school.", "هيكلة المصحف", 15),
]

ws = wb.active
ws.title = "Glossary"
ws["A1"] = "جدول حصر المصطلحات القرآنية — القرآن الكريم"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:E1")
ws["A2"] = (
    "Built from QUL (qul.tarteel.ai) structural metadata — counts are exact, not estimates. "
    "Companion to community.itqan.dev/d/521; rows below can be pasted into that shared sheet."
)
ws["A2"].font = NOTE_FONT
ws.merge_cells("A2:E2")

headers = ["المصطلح (Term)", "Transliteration", "شرح المعنى (Meaning)", "تصنيف (Category)", "Count in Qur'an"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

r = 5
for term_ar, translit, meaning, category, count in glossary:
    ws.cell(row=r, column=1, value=term_ar).font = ARABIC_FONT
    ws.cell(row=r, column=2, value=translit)
    ws.cell(row=r, column=3, value=meaning)
    ws.cell(row=r, column=4, value=category)
    ws.cell(row=r, column=5, value=count)
    style_body_row(ws, r, len(headers))
    r += 1

autosize(ws, [16, 16, 62, 16, 14])
ws.freeze_panes = "A5"

# ---- Sheet 2: Surahs (all 114) ---------------------------------------------

surah_data = load("quran-metadata-surah-name.json")
ws2 = wb.create_sheet("Surahs")
ws2["A1"] = "Surahs — all 114"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:H1")
ws2["A2"] = "Source: QUL quran-metadata-surah-name.json"
ws2["A2"].font = NOTE_FONT
ws2.merge_cells("A2:H2")

headers2 = ["#", "الاسم العربي", "Name", "Translated Meaning", "Revelation Place", "Revelation Order", "Ayah Count", "Bismillah Before?"]
for i, h in enumerate(headers2, start=1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, len(headers2))

r = 5
for i in range(1, 115):
    rec = surah_data[str(i)]
    ws2.cell(row=r, column=1, value=rec["id"])
    ws2.cell(row=r, column=2, value=rec["name_arabic"]).font = ARABIC_FONT
    ws2.cell(row=r, column=3, value=rec["name_simple"])
    ws2.cell(row=r, column=4, value=rec["name"])
    ws2.cell(row=r, column=5, value=rec["revelation_place"].capitalize())
    ws2.cell(row=r, column=6, value=rec["revelation_order"])
    ws2.cell(row=r, column=7, value=rec["verses_count"])
    ws2.cell(row=r, column=8, value="Yes" if rec["bismillah_pre"] else "No")
    style_body_row(ws2, r, len(headers2))
    r += 1

autosize(ws2, [6, 16, 16, 22, 15, 14, 11, 14])
ws2.freeze_panes = "A5"

# ---- Sheets 3-6: Juz / Hizb / Rub / Manzil ranges --------------------------

def build_range_sheet(name, filename, number_field, title):
    data = load(filename)
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Source: QUL {filename}"
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:D2")

    headers = ["#", "First Ayah", "Last Ayah", "Ayah Count"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(headers))

    r = 5
    for rec in data.values():
        ws.cell(row=r, column=1, value=rec[number_field])
        ws.cell(row=r, column=2, value=rec["first_verse_key"])
        ws.cell(row=r, column=3, value=rec["last_verse_key"])
        ws.cell(row=r, column=4, value=rec["verses_count"])
        style_body_row(ws, r, len(headers))
        r += 1

    autosize(ws, [8, 14, 14, 14])
    ws.freeze_panes = "A5"


build_range_sheet("Juz", "quran-metadata-juz.json", "juz_number", "Juz' — all 30")
build_range_sheet("Hizb", "quran-metadata-hizb.json", "hizb_number", "Hizb — all 60")
build_range_sheet("Rub", "quran-metadata-rub.json", "rub_number", "Rub' al-Hizb — all 240")
build_range_sheet("Manzil", "quran-metadata-manzil.json", "manzil_number", "Manzil — all 7")

# ---- Sheet 7: Ruku (has surah_ruku_number too) -----------------------------

ruku_data = load("quran-metadata-ruku.json")
ws7 = wb.create_sheet("Ruku")
ws7["A1"] = "Ruku' — all 558"
ws7["A1"].font = TITLE_FONT
ws7.merge_cells("A1:E1")
ws7["A2"] = "Source: QUL quran-metadata-ruku.json"
ws7["A2"].font = NOTE_FONT
ws7.merge_cells("A2:E2")

headers7 = ["# (global)", "# (within surah)", "First Ayah", "Last Ayah", "Ayah Count"]
for i, h in enumerate(headers7, start=1):
    ws7.cell(row=4, column=i, value=h)
style_header(ws7, 4, len(headers7))

r = 5
for rec in ruku_data.values():
    ws7.cell(row=r, column=1, value=rec["ruku_number"])
    ws7.cell(row=r, column=2, value=rec["surah_ruku_number"])
    ws7.cell(row=r, column=3, value=rec["first_verse_key"])
    ws7.cell(row=r, column=4, value=rec["last_verse_key"])
    ws7.cell(row=r, column=5, value=rec["verses_count"])
    style_body_row(ws7, r, len(headers7))
    r += 1

autosize(ws7, [12, 16, 12, 12, 12])
ws7.freeze_panes = "A5"

# ---- Sheet 8: Sajda (15) ----------------------------------------------------

sajda_data = load("quran-metadata-sajda.json")
ws8 = wb.create_sheet("Sajda")
ws8["A1"] = "Sajdah (Prostration) Points — all 15"
ws8["A1"].font = TITLE_FONT
ws8.merge_cells("A1:C1")
ws8["A2"] = "Source: QUL quran-metadata-sajda.json (standard Hafs list)"
ws8["A2"].font = NOTE_FONT
ws8.merge_cells("A2:C2")

headers8 = ["#", "Ayah", "Type"]
for i, h in enumerate(headers8, start=1):
    ws8.cell(row=4, column=i, value=h)
style_header(ws8, 4, len(headers8))

r = 5
for rec in sajda_data.values():
    ws8.cell(row=r, column=1, value=rec["sajdah_number"])
    ws8.cell(row=r, column=2, value=rec["verse_key"])
    ws8.cell(row=r, column=3, value=rec["sajdah_type"].capitalize())
    style_body_row(ws8, r, len(headers8))
    r += 1

autosize(ws8, [8, 14, 14])
ws8.freeze_panes = "A5"

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"wrote {OUT}")
