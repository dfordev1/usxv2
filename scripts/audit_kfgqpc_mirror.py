#!/usr/bin/env python3
"""Cross-check Talha bin Bashir's aligned workbook against a pinned KFGQPC mirror."""

import argparse
import collections
import gzip
import hashlib
import json
import subprocess
from pathlib import Path

from openpyxl import load_workbook
from rapidfuzz.distance import Levenshtein

from import_eight_riwayah_workbook import normalize_letters

ROOT = Path(__file__).resolve().parents[1]
MIRROR_FILES = {
    "hafs-kufi": "hafs/data/hafsData_v18.json",
    "shubah-kfqc": "shouba/data/ShoubaData08.json",
    "warsh-kfqc": "warsh/data/warshData_v10.json",
    "qalon-kfqc": "qaloon/data/QaloonData_v10.json",
    "bazzi-kfqc": "bazzi/data/BazziData_v07.json",
    "douri-kfqc": "doori/data/DooriData_v09.json",
    "qunbul-kfqc": "qumbul/data/QumbulData_v07.json",
    "sousi-kfqc": "soosi/data/SoosiData09.json",
}
BASMALA = normalize_letters("بسم الله الرحمن الرحيم")


def sha256(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def mirror_commit(path):
    try:
        return subprocess.check_output(
            ["git", "-C", str(path), "rev-parse", "HEAD"], text=True
        ).strip()
    except (OSError, subprocess.CalledProcessError):
        return None


def source_surah(record):
    return int(record.get("sura_no", record.get("sora")))


def classify_operation(surah, source_text, mirror_text):
    if surah == 1 and source_text == BASMALA and not mirror_text:
        return "boundary-convention"
    if source_text == "ء" and not mirror_text:
        return "encoding-presentation"
    return "substantive-unresolved"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("mirror", type=Path)
    parser.add_argument("--output", type=Path, default=ROOT / "data" / "alignments" / "eight-riwayah-mirror-audit-v1.json")
    args = parser.parse_args()

    with gzip.open(ROOT / "data" / "alignments" / "eight-riwayah-slots-v1.json.gz", "rt", encoding="utf8") as source:
        slots = json.load(source)["slots"]
    row_surahs = [int(slot["canonicalLocations"][0].split(":", 1)[0]) for slot in slots]

    sheet = load_workbook(args.workbook, read_only=True, data_only=True).active
    columns = list(zip(*list(sheet.iter_rows(min_row=2, min_col=2, max_col=9, values_only=True))))
    results = []
    for (tradition, relative_path), column in zip(MIRROR_FILES.items(), columns):
        path = args.mirror / relative_path
        records = json.loads(path.read_text(encoding="utf-8-sig"))
        workbook_surahs = collections.defaultdict(str)
        mirror_surahs = collections.defaultdict(str)
        for surah, value in zip(row_surahs, column):
            workbook_surahs[surah] += str(value or "")
        for record in records:
            mirror_surahs[source_surah(record)] += record["aya_text"]

        differences = []
        for surah in range(1, 115):
            left = normalize_letters(workbook_surahs[surah])
            right = normalize_letters(mirror_surahs[surah])
            for opcode in Levenshtein.opcodes(left, right):
                if opcode.tag == "equal":
                    continue
                source_text = left[opcode.src_start:opcode.src_end]
                mirror_text = right[opcode.dest_start:opcode.dest_end]
                differences.append({
                    "surah": surah,
                    "operation": opcode.tag,
                    "workbook": source_text,
                    "mirror": mirror_text,
                    "classification": classify_operation(surah, source_text, mirror_text),
                })
        counts = collections.Counter(item["classification"] for item in differences)
        results.append({
            "tradition": tradition,
            "path": relative_path,
            "sha256": sha256(path),
            "verseRecords": len(records),
            "normalizedExact": not differences,
            "differenceCounts": dict(sorted(counts.items())),
            "differences": differences,
        })

    report = {
        "format": "qusx-kfgqpc-mirror-audit",
        "version": "1.0.0",
        "status": "cross-source-corroboration-not-scholarly-certification",
        "workbookSha256": sha256(args.workbook),
        "mirror": {
            "repository": "https://github.com/thetruetruth/quran-data-kfgqpc",
            "commit": mirror_commit(args.mirror),
            "licenseStatus": "no-license-file",
        },
        "normalization": "QUSX letter normalization; differences retained and classified after normalization.",
        "results": results,
        "unresolvedSubstantiveDifferences": sum(
            item["differenceCounts"].get("substantive-unresolved", 0) for item in results
        ),
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf8")
    print(json.dumps(report, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
