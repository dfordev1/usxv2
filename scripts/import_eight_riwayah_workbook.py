#!/usr/bin/env python3
"""Import the Talha bin Bashir eight-riwayah workbook as non-normative QUSX slots."""

import argparse
import difflib
import gzip
import hashlib
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXPECTED_HEADERS = ["التسلسل", "حفص", "شعبة", "ورش", "قالون", "البزي", "الدوري", "قنبل", "السوسي", "اختلاف الوصل والفصل"]
TRADITIONS = ["hafs-kufi", "shubah-kfqc", "warsh-kfqc", "qalon-kfqc", "bazzi-kfqc", "douri-kfqc", "qunbul-kfqc", "sousi-kfqc"]
ARABIC_NUMBER = re.compile(r"^[\u0660-\u0669\u06f0-\u06f9]+$")
LETTER_FOLD = {
    0x0671: chr(0x627), 0x0623: chr(0x627), 0x0625: chr(0x627), 0x0622: chr(0x627),
    0x0649: chr(0x64A), 0x06D2: chr(0x64A), 0x0624: chr(0x648), 0x0626: chr(0x64A),
}


def normalize_letters(value):
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(
        char for char in text
        if not (0x0610 <= ord(char) <= 0x061A or 0x064B <= ord(char) <= 0x065F or ord(char) == 0x0670 or 0x06D6 <= ord(char) <= 0x06ED)
    )
    text = text.replace(chr(0x0640), "").replace(chr(0x06DE), "")
    return "".join(char for char in text if "ARABIC" in unicodedata.name(char, "") and unicodedata.category(char).startswith("L")).translate(LETTER_FOLD)


def classify(values, split_join_flag):
    if split_join_flag or any(not value or len(value.split()) > 1 for value in values):
        return "split-join"
    if len({unicodedata.normalize("NFC", value) for value in values}) == 1:
        return "identical"
    if len({normalize_letters(value) for value in values}) == 1:
        return "orthographic-presentation"
    return "substantive-candidate"


def load_qusx_words():
    source = json.loads((ROOT / "data" / "raw" / "uthmani.json").read_text(encoding="utf8"))
    return [
        record for record in sorted(source.values(), key=lambda item: item["id"])
        if not ARABIC_NUMBER.fullmatch(record["text"].strip())
    ]


def load_rows(source_path):
    worksheet = load_workbook(source_path, read_only=True, data_only=True).worksheets[0]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(iterator)]
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Unexpected workbook headers: {headers}")
    rows = list(iterator)
    for expected, row in enumerate(rows, 1):
        if int(row[0]) != expected:
            raise ValueError(f"Workbook sequence breaks at row {expected + 1}: {row[0]}")
    return worksheet.title, rows


def build_location_map(words, rows):
    left = [normalize_letters(record["text"]) for record in words]
    right = [normalize_letters(row[1]) for row in rows]
    matcher = difflib.SequenceMatcher(None, left, right, autojunk=False)
    locations = [[] for _ in rows]
    alignment_blocks = []
    for tag, left_start, left_end, right_start, right_end in matcher.get_opcodes():
        left_size, right_size = left_end - left_start, right_end - right_start
        if tag == "equal" or left_size == right_size:
            for offset in range(right_size):
                locations[right_start + offset] = [words[left_start + offset]["location"]]
        elif right_size == 1:
            locations[right_start] = [record["location"] for record in words[left_start:left_end]]
        elif left_size == 1:
            for index in range(right_start, right_end):
                locations[index] = [words[left_start]["location"]]
        else:
            raise ValueError(f"Unsupported alignment block: {(tag, left_start, left_end, right_start, right_end)}")
        if tag != "equal":
            alignment_blocks.append({
                "kind": tag,
                "qusxLocations": [record["location"] for record in words[left_start:left_end]],
                "workbookSequences": [int(row[0]) for row in rows[right_start:right_end]],
            })
    if any(not value for value in locations):
        raise ValueError("At least one workbook row was not mapped to a QUSX location")
    return locations, alignment_blocks, matcher.ratio()


def write_gzip_json(path, value):
    payload = json.dumps(value, ensure_ascii=False, separators=(",", ":")).encode("utf8")
    with path.open("wb") as raw:
        with gzip.GzipFile(filename="", mode="wb", fileobj=raw, mtime=0) as compressed:
            compressed.write(payload)
    return len(payload)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    args = parser.parse_args()
    source_path = args.source.resolve()
    source_hash = hashlib.sha256(source_path.read_bytes()).hexdigest()
    sheet_name, rows = load_rows(source_path)
    words = load_qusx_words()
    if len(rows) != 77432 or len(words) != 77432:
        raise ValueError(f"Expected 77,432 workbook/QUSX units, found {len(rows)}/{len(words)}")
    locations, alignment_blocks, ratio = build_location_map(words, rows)

    counts = {key: 0 for key in ["identical", "orthographic-presentation", "substantive-candidate", "split-join"]}
    slots, candidates = [], []
    repeated_locations = {}
    for mapped in locations:
        for location in mapped:
            repeated_locations[location] = repeated_locations.get(location, 0) + 1
    seen_locations = {}

    for index, (row, mapped) in enumerate(zip(rows, locations), 1):
        values = [str(value or "").strip() for value in row[1:9]]
        split_join = row[9] not in (None, "", 0, "0")
        category = classify(values, split_join)
        counts[category] += 1
        segment = None
        if len(mapped) == 1 and repeated_locations[mapped[0]] > 1:
            seen_locations[mapped[0]] = seen_locations.get(mapped[0], 0) + 1
            segment = {"index": seen_locations[mapped[0]], "total": repeated_locations[mapped[0]]}
        slot = {
            "id": f"qusx:slot:global:{index:06d}",
            "sequence": index,
            "canonicalLocations": mapped,
            "canonicalSegment": segment,
            "classification": category,
            "splitJoinFlag": split_join,
            "readings": {tradition: value.split() if value else [] for tradition, value in zip(TRADITIONS, values)},
            "provenance": {"sheet": sheet_name, "row": index + 1},
        }
        slots.append(slot)
        if category in ("substantive-candidate", "split-join"):
            candidates.append(slot)

    manifest = {
        "format": "qusx-eight-riwayah-slots",
        "version": "1.0.0",
        "status": "source-derived-candidate-review-required",
        "source": {
            "title": "ملف كلمات الروايات",
            "compiler": "Talha bin Bashir",
            "url": "https://docs.google.com/spreadsheets/d/1iGgb9Q-pYmDHKggTGuErcUJc6Yal8w3q/edit",
            "sha256": source_hash,
            "redistributionStatus": "permission-not-yet-documented-in-repository",
            "method": "Text extracted from eight King Fahd Complex mushaf/font editions and manually aligned to the 77,432-row Hafs sequence.",
        },
        "canonicalTradition": "hafs-kufi",
        "traditions": TRADITIONS,
        "slotCount": len(slots),
        "classificationCounts": counts,
        "qusxHafsAlignment": {"ratio": ratio, "nonEqualBlocks": alignment_blocks},
    }
    full_document = {**manifest, "slots": slots}
    candidate_document = {**manifest, "candidateCount": len(candidates), "slots": candidates}
    output_dir = ROOT / "data" / "alignments"
    full_size = write_gzip_json(output_dir / "eight-riwayah-slots-v1.json.gz", full_document)
    candidate_size = write_gzip_json(output_dir / "eight-riwayah-review-candidates-v1.json.gz", candidate_document)
    (output_dir / "eight-riwayah-manifest-v1.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf8")
    print(json.dumps({"slots": len(slots), "candidates": len(candidates), "counts": counts, "sourceSha256": source_hash, "uncompressedBytes": {"full": full_size, "candidates": candidate_size}}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
